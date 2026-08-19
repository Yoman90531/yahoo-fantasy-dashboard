from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from urllib.request import Request, urlopen

from sqlalchemy.orm import Session

from app.models.adp import AdpEntry, AdpSnapshot
from app.services.keepers import normalize_player_name


@dataclass(frozen=True)
class AdpRecord:
    rank: int
    player_name: str
    position: str | None
    nfl_team: str | None
    average_adp: float | None


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _number(value: object) -> float | None:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _clean_position(value: object) -> str | None:
    cleaned = re.sub(r"\d+$", "", str(value or "").strip().upper())
    if cleaned in {"D/ST", "DEFENSE"}:
        return "DST"
    return cleaned or None


def _player_and_team(player_value: object, team_value: object = None) -> tuple[str, str | None]:
    raw = re.sub(r"\s+", " ", str(player_value or "")).strip()
    explicit_team = str(team_value or "").strip().upper() or None
    parenthetical = re.match(r"^(.*?)\s+\(([A-Z]{2,3})\)(?:\s*\(\d+\))?$", raw)
    if parenthetical:
        return parenthetical.group(1).strip(), explicit_team or parenthetical.group(2)

    trailing = re.match(r"^(.*?)\s+([A-Z]{2,3})\s+\(\d+\)$", raw)
    if trailing:
        name = re.sub(r"\s+[A-Z]\.\s+\S+$", "", trailing.group(1)).strip()
        return name, explicit_team or trailing.group(2)
    return raw, explicit_team


def records_from_rows(headers: Iterable[object], rows: Iterable[Iterable[object]]) -> list[AdpRecord]:
    normalized_headers = [_normalize_header(header) for header in headers]

    def index_of(*candidates: str) -> int | None:
        for candidate in candidates:
            if candidate in normalized_headers:
                return normalized_headers.index(candidate)
        for index, header in enumerate(normalized_headers):
            if any(candidate in header for candidate in candidates):
                return index
        return None

    rank_index = index_of("rank", "rk")
    player_index = index_of("player team bye", "player name", "player")
    position_index = index_of("pos", "position")
    team_index = next(
        (normalized_headers.index(header) for header in ("team", "nfl team") if header in normalized_headers),
        None,
    )
    average_index = index_of("avg", "average", "consensus adp")
    if rank_index is None or player_index is None:
        raise ValueError("ADP data must contain Rank and Player columns")

    records: list[AdpRecord] = []
    seen_ranks: set[int] = set()
    for raw_row in rows:
        row = list(raw_row)
        if rank_index >= len(row) or player_index >= len(row):
            continue
        raw_rank = _number(row[rank_index])
        if raw_rank is None or raw_rank < 1:
            continue
        rank = int(raw_rank)
        if rank in seen_ranks:
            continue
        team_value = row[team_index] if team_index is not None and team_index < len(row) else None
        player_name, nfl_team = _player_and_team(row[player_index], team_value)
        if not player_name:
            continue
        position = (
            _clean_position(row[position_index])
            if position_index is not None and position_index < len(row)
            else None
        )
        average = (
            _number(row[average_index])
            if average_index is not None and average_index < len(row)
            else None
        )
        seen_ranks.add(rank)
        records.append(AdpRecord(rank, player_name, position, nfl_team, average))

    return sorted(records, key=lambda record: record.rank)


def read_csv_records(path: Path) -> list[AdpRecord]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return []
    return records_from_rows(rows[0], rows[1:])


def read_excel_records(path: Path) -> list[AdpRecord]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Install openpyxl to import .xlsx ADP files") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    elif suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("Install xlrd to import .xls ADP files") from exc
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
    else:
        raise ValueError(f"Unsupported spreadsheet type: {suffix}")
    if not rows:
        return []
    return records_from_rows(rows[0], rows[1:])


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell_parts: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in {"th", "td"} and self._row is not None:
            self._cell_parts = []

    def handle_data(self, data: str) -> None:
        if self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self._cell_parts is not None and self._row is not None:
            self._row.append(re.sub(r"\s+", " ", " ".join(self._cell_parts)).strip())
            self._cell_parts = None
        elif tag == "tr" and self._row is not None and self._table is not None:
            if self._row:
                self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            if self._table:
                self.tables.append(self._table)
            self._table = None


def records_from_html(html: str) -> list[AdpRecord]:
    marker = "window.FP.reportConfig ="
    marker_index = html.find(marker)
    if marker_index >= 0:
        payload_start = marker_index + len(marker)
        payload = html[payload_start:].lstrip()
        try:
            config, _ = json.JSONDecoder().raw_decode(payload)
            embedded_rows = config.get("table", {}).get("rows", [])
            records = []
            for row in embedded_rows:
                rank_value = _number(row.get("rank"))
                player = row.get("player") or {}
                player_name = str(player.get("name") or "").strip()
                if rank_value is None or not player_name:
                    continue
                raw_team = str(player.get("team") or "").strip()
                team_match = re.match(r"^([A-Z]{2,3})(?:\s+\(\d+\))?$", raw_team)
                records.append(
                    AdpRecord(
                        rank=int(rank_value),
                        player_name=player_name,
                        position=_clean_position(row.get("pos")),
                        nfl_team=team_match.group(1) if team_match else None,
                        average_adp=_number(row.get("avg")),
                    )
                )
            if records:
                return sorted(records, key=lambda record: record.rank)
        except (json.JSONDecodeError, TypeError, AttributeError):
            pass

    parser = _TableParser()
    parser.feed(html)
    for table in parser.tables:
        for header_index, row in enumerate(table[:5]):
            normalized = {_normalize_header(cell) for cell in row}
            has_rank = "rank" in normalized or "rk" in normalized
            has_player = any("player" in cell for cell in normalized)
            if has_rank and has_player:
                records = records_from_rows(row, table[header_index + 1 :])
                if records:
                    return records
    raise ValueError("Could not find a FantasyPros ADP table in the downloaded page")


def fetch_html_records(url: str, timeout: int = 30) -> list[AdpRecord]:
    request = Request(
        url,
        headers={
            "User-Agent": "GARYS-Fantasy-Dashboard/1.0 keeper-adp-import",
            "Accept": "text/html,application/xhtml+xml",
        },
    )
    with urlopen(request, timeout=timeout) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        html = response.read().decode(charset, errors="replace")
    return records_from_html(html)


def read_records(path: Path) -> list[AdpRecord]:
    if path.suffix.lower() == ".csv":
        return read_csv_records(path)
    return read_excel_records(path)


def store_snapshot(
    db: Session,
    records: list[AdpRecord],
    *,
    season: int,
    source: str,
    source_url: str | None,
    scoring_format: str,
    league_size: int,
    lock: bool = True,
) -> AdpSnapshot:
    if not records:
        raise ValueError("Cannot store an empty ADP snapshot")
    ranks = [record.rank for record in records]
    if len(ranks) != len(set(ranks)):
        raise ValueError("ADP ranks must be unique")
    if lock:
        db.query(AdpSnapshot).filter(
            AdpSnapshot.season == season,
            AdpSnapshot.scoring_format == scoring_format,
        ).update({"is_locked": False}, synchronize_session=False)

    snapshot = AdpSnapshot(
        season=season,
        source=source,
        source_url=source_url,
        scoring_format=scoring_format,
        league_size=league_size,
        is_locked=lock,
    )
    db.add(snapshot)
    db.flush()
    db.add_all(
        [
            AdpEntry(
                snapshot_id=snapshot.id,
                rank=record.rank,
                player_name=record.player_name,
                normalized_name=normalize_player_name(record.player_name),
                position=record.position,
                nfl_team=record.nfl_team,
                average_adp=record.average_adp,
            )
            for record in records
        ]
    )
    db.commit()
    db.refresh(snapshot)
    return snapshot
